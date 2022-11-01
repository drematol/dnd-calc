from math import ceil
import openpyxl as xl


def main():
    wb = xl.load_workbook("character.xlsx")
    sheet = wb["Sheet1"]

    # inputting the non derived statistics
    name = sheet.cell(1, 2).value

    def spliter(s):
        if " / " in s:
            return s.split(" / ")
        elif "/" in s:
            return s.split("/")
        elif " \\ " in s:
            return s.split(" \\ ")
        elif "\\" in s:
            return s.split("\\")
        else:
            return s.split()

    player_class = spliter(sheet.cell(2, 2).value.lower())
    level = spliter(str(sheet.cell(3, 2).value))

    if len(level) > len(player_class):
        for i in range(len(level) - len(player_class)):
            level.pop(-1)
    if len(level) < len(player_class):
        for i in range(len(player_class) - len(level)):
            level.append("0")

    race = sheet.cell(4, 2).value.lower()

    try:
        armor = sheet.cell(5, 2).value.lower()
    except AttributeError:
        armor = "none"

    strength = sheet.cell(6, 2).value // 2 - 5
    dexterity = sheet.cell(7, 2).value // 2 - 5
    constitution = sheet.cell(8, 2).value // 2 - 5
    intelligence = sheet.cell(9, 2).value // 2 - 5
    wisdom = sheet.cell(10, 2).value // 2 - 5
    charisma = sheet.cell(11, 2).value // 2 - 5

    # some dictionaries to define traits for our inputs
    classes = {
        "barbarian": [12, "Null"],
        "fighter": [10, "Null"],
        "paladin": [10, "cha"],
        "ranger": [10, "wis"],
        "artificer": [8, "int"],
        "bard": [8, "cha"],
        "cleric": [8, "wis"],
        "druid": [8, "wis"],
        "monk": [8, "Null"],
        "rouge": [8, "Null"],
        "warlock": [8, "cha"],
        "sorcerer": [6, "cha"],
        "wizard": [6, "int"]
    }
    races = {
        "air genasi": 35,
        "dhampir": 35,
        "satyr": 35,
        "wood elf": 35,
        "dwarf": 25,
        "gnome": 25,
        "halfling": 25,
        "aarakocra": 25,
        "aven": 25,
        "siren": 25
    }
    armores = {
        "padded": ["light", 11],
        "leather": ["light", 11],
        "studded leather": ["light", 12],
        "studded": ["light",  12],
        "hide": ["medium", 12],
        "chain shirt": ["medium", 13],
        "scale mail": ["medium", 14],
        "scale": ["medium", 14],
        "breastplate": ["medium", 14],
        "breast plate": ["medium", 14],
        "half plate": ["medium", 15],
        "ring mail": ["heavy", 14],
        "ring": ["heavy", 14],
        "chain mail": ["heavy", 16],
        "splint": ["heavy", 17],
        "plate": ["heavy", 18]
    }

    # calculating derived statistics
    spell_casting = []
    hp = classes.get(player_class[0], [0, "none"])[0] + constitution
    hit_dice_str = ""
    for idx, i in enumerate(player_class):
        hit_dice = classes.get(i, [0, "none"])[0]
        if idx == 0:
            hp += (hit_dice // 2 + constitution) * (int(level[idx])-1)
        if idx != 0:
            hp += (hit_dice // 2 + constitution) * int(level[idx])
        spell_casting.append(classes.get(i, [0, "none"])[1])
        hit_dice_str += level[idx] + "d" + str(hit_dice) + " "
    total_level = 0
    for i in level:
        total_level += int(i)
    proficiency_bonus = ceil(total_level/4) + 1
    category = armores.get(armor, ["light", 10])[0]
    ac_bonus = armores.get(armor, ["light", 10])[1]
    ac = 10
    if category == "light":
        ac = ac_bonus + dexterity
    if category == "medium":
        ac = ac_bonus + min(dexterity, 2)
    if category == "heavy":
        ac = ac_bonus
    speed = races.get(race, 30)

    # out putting derived statistics
    sheet.cell(1, 4).value = "hp"
    sheet.cell(1, 5).value = hp
    sheet.cell(2, 4).value = "hit dice"
    sheet.cell(2, 5).value = hit_dice_str
    sheet.cell(3, 4).value = "prof bon"
    sheet.cell(3, 5).value = proficiency_bonus
    sheet.cell(4, 4).value = "ac"
    sheet.cell(4, 5).value = ac
    sheet.cell(5, 4).value = "speed"
    sheet.cell(5, 5).value = speed
    sheet.cell(6, 4).value = "init"
    sheet.cell(6, 5).value = dexterity

    # adding modifiers to all 6 ability scores
    def modifier(val, score):
        if score >= 0:
            return str(val) + " / +" + str(score)
        if score < 0:
            return str(val) + " / " + str(score)
    sheet.cell(6, 2).value = modifier(sheet.cell(6, 2).value, strength)
    sheet.cell(7, 2).value = modifier(sheet.cell(7, 2).value, dexterity)
    sheet.cell(8, 2).value = modifier(sheet.cell(8, 2).value, constitution)
    sheet.cell(9, 2).value = modifier(sheet.cell(9, 2).value, intelligence)
    sheet.cell(10, 2).value = modifier(sheet.cell(10, 2).value, wisdom)
    sheet.cell(11, 2).value = modifier(sheet.cell(11, 2).value, charisma)

    # spell casting if relevant
    i = 7
    if "int" in spell_casting:
        sheet.cell(i, 4).value = "ability"
        sheet.cell(i, 5).value = "int"
        i += 1
        sheet.cell(i, 4).value = "spell dc"
        sheet.cell(i, 5).value = 8 + proficiency_bonus + intelligence
        i += 1
        sheet.cell(i, 4).value = "spell bon"
        sheet.cell(i, 5).value = proficiency_bonus + intelligence
        i += 1
    if "wis" in spell_casting:
        sheet.cell(i, 4).value = "ability"
        sheet.cell(i, 5).value = "wis"
        i += 1
        sheet.cell(i, 4).value = "spell dc"
        sheet.cell(i, 5).value = 8 + proficiency_bonus + wisdom
        i += 1
        sheet.cell(i, 4).value = "spell bon"
        sheet.cell(i, 5).value = proficiency_bonus + wisdom
        i += 1
    if "cha" in spell_casting:
        sheet.cell(i, 4).value = "ability"
        sheet.cell(i, 5).value = "cha"
        i += 1
        sheet.cell(i, 4).value = "spell dc"
        sheet.cell(i, 5).value = 8 + proficiency_bonus + charisma
        i += 1
        sheet.cell(i, 4).value = "spell bon"
        sheet.cell(i, 5).value = proficiency_bonus + charisma
        i += 1

    # and then we save the new spreadsheet
    try:
        name += ".xlsx"
        wb.save(name)
    except TypeError:
        wb.save("new character.xlsx")

    # when you format your level as x/x/x it assumes it's a date and will forever format that cell as a date
    # witch will mark anything being saved into that cell as a date
    # since this is a "spreadsheet" thing and not a "my code" thing I cannot think of an easy fix
    # and don't think the problem is relevant enough to warrant the time it would take to fix it


if __name__ == '__main__':
    main()

# I understand that this code is uncleanly formatted as this is my first big project
# the point of the project is to practice automation with python
# if you notice any big problems please let me know
